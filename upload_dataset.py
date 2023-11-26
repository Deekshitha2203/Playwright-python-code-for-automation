import os
import openpyxl
from playwright.sync_api import sync_playwright
username = 'atharrafiqi+62@datoin.com'
password = 'Sphi@1234'
s = "//p[normalize-space()='Start Training']"
with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    context = browser.new_context ()
    page = context.new_page()
    def login(page):
        page.goto("https://pre-app.datoin.com")
        un = page.locator('#username')
        un.fill(username)
        page.wait_for_selector("#username")
        pas = page.get_by_label('Password')
        pas.fill(password)
        page.get_by_role("button",name='Sign In').click()
        selector = "//p[normalize-space()='Acquire Data']"
        s = "//p[normalize-space()='Start Training']"
        h = "//p[normalize-space()='Evaluate Model']"
        try:
            element = page.wait_for_selector(selector, timeout=240000)
            try:
                ele = page.wait_for_selector(s,timeout=240000)
                try:
                    el = page.wait_for_selector(h,timeout=240000)
                    page.screenshot(path='screenshot.png')
                    print("Login successful!")
                except Exception as e:
                    print('Element is not present on the page or did not appear within the timeout.')
            except Exception as e:
                print('Element is not present on the page or did not appear within the timeout.')
        except Exception as e:
            print('Element is not present on the page or did not appear within the timeout.')
    login(page)

    
    def upload_dataset(page):
        da = "//div[contains(text(),'Datasets')]"
        page.click(da)
        page.screenshot(path='screen.png')
        ds = '//div[@class="admin-panel-title"]'
        ch = '//p[normalize-space()="Upload Dataset"]'
        fold = '//div[contains(text(),"Folders")]'
        try:
            ob = page.wait_for_selector(ds,timeout=240000)
            try:
                oob = page.wait_for_selector(ch,timeout=240000)
                try:
                    ooob = page.wait_for_selector(fold,timeout=240000)
                    print("Datasets opened successfully!")
                    page.screenshot(path='datashot.png')
                except Exception as e:
                    print('Element is not present on the page or did not appear within the timeout.')
            except Exception as e:
                    print('Element is not present on the page or did not appear within the timeout.')
        except Exception as e:
            print('Element is not present on the page or did not appear within the timeout.')
        directory_path = " " #enter the path to the directory
        column_letter = 'A'
        wb = openpyxl.load_workbook('data1.xlsx',data_only=True)
        sheet = wb['Sheet1']
        column = sheet[column_letter]
        total_rows =  sum(1 for cell in column if cell.value is not None)
        for filename in os.listdir(directory_path):
            is_present = False
            print(total_rows)
            for row in range(1,total_rows+1):
                dataset = sheet.cell(row=row, column=1).value
                if filename == dataset:
                    is_present = True
                    break
            print(is_present)
            if not is_present:
                file_path = os.path.join(directory_path, filename)
                upload = "//p[normalize-space()='Upload Dataset']"
                element = page.wait_for_selector(upload, state='visible')
                element.click()
                file_input = page.locator('input[type="file"]')
                file_input.set_input_files(file_path)
                drop_area = page.locator("//div[@class='filepond--drop-label']")
                drop_area_bbox = drop_area.bounding_box()
                drop_x = drop_area_bbox['x'] + drop_area_bbox['width'] / 2
                drop_y = drop_area_bbox['y'] + drop_area_bbox['height'] / 2
                page.mouse.move(drop_x, drop_y)
                page.mouse.down()
                page.mouse.move(drop_x, drop_y)
                page.mouse.up()
                upload_button_selector = "//button[normalize-space()='Upload']"
                upload_button = page.wait_for_selector(upload_button_selector, state='visible', timeout=5000)
                upload_button.click()
                pau = "//button[normalize-space()='Pause All']"
                try:
                    pau = page.wait_for_selector(pau, timeout=240000)
                except Exception as e:
                    print("Problem in uploading!")
                upload_complete_selector = "//div[@class='button-cancel']"
                try:
                    exe = page.wait_for_selector(upload_complete_selector, timeout=240000)
                except Exception as e:
                    print("Problem in uploading!")
                filename=filename.replace('.csv','')
                done_selector = '//div[@title='+"'"+filename+"']"
                try:
                    ex = page.wait_for_selector(done_selector, timeout=240000)
                    print(filename+" uploaded successfully!")
                except Exception as e:
                    print(filename+" is not uploaded!")
                start_row = total_rows + 1
                newvalue = filename+'.csv'
                print(newvalue)
                sheet.cell(row=start_row, column=1, value=newvalue)
                sheet.cell(row=start_row, column=2, value='uploaded')
                total_rows += 1
        wb.close()
        wb.save('dsheet.xlsx')
    upload_dataset(page)

    
    def create_new_application(page):
        new = "https://pre-app.datoin.com/4ed52767-3689-4951-85c2-0bab8b560222/7ef5d230-8dd0-471a-b180-bc83ef88efd6/apps"
        page.goto(new,wait_until="load")
        page.keyboard.press("Escape")        # apps = "//div[@title='My Apps']"
        # page.wait_for_selector(apps,timeout=240000)
        # doapp = page.click(apps)
        newapp = "//button[normalize-space()='New Application']"
        clap = page.click(newapp)
        cross = '//*[@id="___reactour"]/div[4]/div/button'
        try:
            page.wait_for_selector(cross,timeout=240000)
            page.click(cross)
        except Exception as e:
            print("Not done!")
        build = '//*[@id="scrollableDiv"]/div/div/div[2]/div/div[2]/div'
        biu = page.click(build)
        textbox = "(//input[@id='tfid-0-0'])[1]"
        fil = page.locator(textbox).fill('Iris')
        ok = "//button[normalize-space()='Start Build »']"
        try:
            buid = page.wait_for_selector(ok,timeout=240000)
        except Exception as e:
            print("Not clicked!")
        oko = page.click(ok)
        startnew = "//button[normalize-space()='Start New Experiment']"
        try:
            wit = page.wait_for_selector(startnew,timeout=240000)
        except Exception as e:
            print("Not found!")
        start = page.click(startnew)
        newex = "//h2[normalize-space()='New Experiment']"
        try:
            witho = page.wait_for_selector(newex,timeout=240000)
        except Exception as e:
            print("Not found!")
        newdata = "//button[normalize-space()='Choose dataset']"
        newd = page.click(newdata)
        plus = "//div[text()='isr' and @class='browser_item_name']/../following-sibling::div//button"
        irs = "//div[@class='browser_item_name'][normalize-space()='isr']"
        try:
            clic = page.wait_for_selector(irs,timeout=240000)
            witty = page.wait_for_selector(plus,timeout=240000)
            witty.click()
            page.screenshot(path="dataset2.png")
        except Exception as e:
            print("Not found!")
        target = "//button[normalize-space()='Choose target']"
        targ = page.click(target)
        tarsel = "//div[@class=' css-tlfecz-indicatorContainer']//*[name()='svg']"
        ast = page.click(tarsel)
        #print(ast)
        te = "//div//input"
        ex = page.locator(te)
        ex.type("Species")
        page.wait_for_timeout(500)
        ex.press('Enter')
        startsear = "//button[normalize-space()='Start Experiment']"
        try:
            page.wait_for_selector(startsear,timeout=240000)
            star = page.click(startsear)
        except Exception as e:
            print("Target not found!")
        okay = '//*[@id="___reactour"]/div[4]/div/div[2]/button[2]/span/button'
        try:
            clok = page.wait_for_selector(okay,timeout=240000)
            page.click(okay)
        except Exception as e:
            print("Done!")
        page.close()
        # use = "//button[normalize-space()='A']"
        # page.click(use)
        # logout = "//button[normalize-space()='Log out']"
        # page.click(logout)

    def result_page(page):
        count = 0
        wb = openpyxl.load_workbook('datasheet-5.xlsx',data_only=True)
        sheet = wb['sheet']
        for row in range(2,11):
            if sheet['F'+str(row)].value is None or sheet['F'+str(row)].value == 'Running':
                Name = sheet['B'+str(row)].value
                jobid = sheet['E'+str(row)].value
                status = sheet['F'+str(row)].value
                print(jobid)
                page.goto(jobid,wait_until="load")
                success = "//button[normalize-space()='View Report']"
                failure = "//button[normalize-space()='View Error']"
                sne = "//i[@class='fa fa-calendar']"
                page.wait_for_selector(sne,timeout=24000)
                done_v = page.is_visible(success)
                print(done_v)
                not_done = page.is_visible(failure)
                print(not_done)
                if done_v:
                    sheet['F'+str(row)].value = 'Successful'
                elif not_done:
                    sheet['F'+str(row)].value = 'Failed'
                else:
                    sheet['F'+str(row)].value = 'Running'
                    count = count+1
        print(count)
        wb.close()
        if count < 2:
                create_new_application(page)
        else:
            print("2 applications are already running. Thus wait till they complete.")
            wb.save('sheet1.xlsx')

    #result_page(page)
